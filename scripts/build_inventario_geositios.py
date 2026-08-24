"""
Procesa el KMZ y XLSX oficial del Inventario Nacional de Geositios SERNAGEOMIN 2024
→ GeoJSON cargable en el visor.

Entrada:
- docs/mapas/chile/inventario_geositios/libro-geositios-de-chile-sernageomin2024.kmz
- docs/mapas/chile/inventario_geositios/libro_geositios_de_chile_sernageomin.xlsx

Salida:
- app/data/geositios_inventario_nacional.geojson (puntos con todos los atributos)
- docs/biblioteca/_db/inventario_geositios.json (mismo dato, indexable)
"""

import zipfile
import json
import io
from pathlib import Path
from xml.etree import ElementTree as ET
import sys

# En Windows, Python solo usa UTF-8 en consola interactiva (PEP 528). Al
# redirigir la salida a un archivo cae a cp1252 y cualquier print() con
# flechas o checks lanza UnicodeEncodeError, abortando el script a medio
# correr. Esto lo fuerza a UTF-8 siempre.
for _s in (sys.stdout, sys.stderr):
    if hasattr(_s, "reconfigure"):
        _s.reconfigure(encoding="utf-8")


ROOT = Path(__file__).resolve().parent.parent
KMZ = ROOT / "docs/mapas/chile/inventario_geositios/libro-geositios-de-chile-sernageomin2024.kmz"
XLSX = ROOT / "docs/mapas/chile/inventario_geositios/libro_geositios_de_chile_sernageomin.xlsx"
OUT_GEOJSON = ROOT / "app/data/geositios_inventario_nacional.geojson"
OUT_DB = ROOT / "docs/biblioteca/_db/inventario_geositios.json"

KML_NS = {"k": "http://www.opengis.net/kml/2.2"}


def parse_kmz(path):
    """Extrae el KML del KMZ y parsea Placemarks usando lxml (más tolerante)."""
    from lxml import etree as LET
    with zipfile.ZipFile(path) as zf:
        kml_name = next(n for n in zf.namelist() if n.endswith(".kml"))
        raw = zf.read(kml_name)
    parser = LET.XMLParser(recover=True)
    root = LET.fromstring(raw, parser)
    placemarks = root.findall(".//{http://www.opengis.net/kml/2.2}Placemark")
    print(f"  {len(placemarks)} Placemarks en KMZ")

    features = []
    for pm in placemarks:
        name_el = pm.find("k:name", KML_NS)
        name = (name_el.text or "").strip() if name_el is not None else "(sin nombre)"
        desc_el = pm.find("k:description", KML_NS)
        desc = (desc_el.text or "").strip() if desc_el is not None else ""

        # Extraer atributos de ExtendedData
        attrs = {}
        for data in pm.findall(".//k:Data", KML_NS):
            key = data.get("name", "")
            value_el = data.find("k:value", KML_NS)
            attrs[key] = (value_el.text or "").strip() if value_el is not None else ""
        for sd in pm.findall(".//k:SchemaData/k:SimpleData", KML_NS):
            key = sd.get("name", "")
            attrs[key] = (sd.text or "").strip()

        # Extraer geometría
        geom = None
        # Punto
        coord_el = pm.find(".//k:Point/k:coordinates", KML_NS)
        if coord_el is not None and coord_el.text:
            parts = coord_el.text.strip().split(",")
            try:
                lon, lat = float(parts[0]), float(parts[1])
                geom = {"type": "Point", "coordinates": [lon, lat]}
            except (ValueError, IndexError):
                pass
        # Polígono
        if geom is None:
            coords_el = pm.find(".//k:Polygon//k:outerBoundaryIs//k:LinearRing/k:coordinates", KML_NS)
            if coords_el is not None and coords_el.text:
                ring = []
                for c in coords_el.text.strip().split():
                    p = c.split(",")
                    try:
                        ring.append([float(p[0]), float(p[1])])
                    except (ValueError, IndexError):
                        pass
                if ring:
                    geom = {"type": "Polygon", "coordinates": [ring]}
        # LineString
        if geom is None:
            coords_el = pm.find(".//k:LineString/k:coordinates", KML_NS)
            if coords_el is not None and coords_el.text:
                line = []
                for c in coords_el.text.strip().split():
                    p = c.split(",")
                    try:
                        line.append([float(p[0]), float(p[1])])
                    except (ValueError, IndexError):
                        pass
                if line:
                    geom = {"type": "LineString", "coordinates": line}

        if geom is None:
            continue

        properties = {"nombre": name, "descripcion": desc, **attrs}
        features.append({"type": "Feature", "properties": properties, "geometry": geom})

    return features


def enrich_from_xlsx(features, xlsx_path):
    """Lee el XLSX atributivo y enriquece cada feature por nombre."""
    try:
        import openpyxl
    except ImportError:
        print("  WARN: openpyxl no instalado. pip install openpyxl. Skip enrich.")
        return features
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    print(f"  Hojas en XLSX: {wb.sheetnames}")
    rows_by_name = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [str(c.value) if c.value is not None else "" for c in ws[1]]
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = {h: (str(v) if v is not None else "") for h, v in zip(headers, r)}
            # Heurística: la columna "Nombre" o similar
            name_keys = [k for k in row if k and "ombre" in k.lower()]
            if name_keys:
                key = row[name_keys[0]].strip().lower()
                if key:
                    rows_by_name[key] = row
    enriched = 0
    for feat in features:
        nm = feat["properties"]["nombre"].strip().lower()
        if nm in rows_by_name:
            for k, v in rows_by_name[nm].items():
                if k and v and k not in feat["properties"]:
                    feat["properties"][k] = v
            enriched += 1
    print(f"  Enriquecidos: {enriched}/{len(features)}")
    return features


def main():
    print("Leyendo KMZ del Inventario Nacional…")
    features = parse_kmz(KMZ)

    print(f"\nLeyendo XLSX atributivo…")
    features = enrich_from_xlsx(features, XLSX)

    # Estadísticas
    from collections import Counter
    geom_types = Counter(f["geometry"]["type"] for f in features)
    regions = Counter()
    for f in features:
        for k, v in f["properties"].items():
            if "egion" in k.lower() and v:
                regions[v] += 1
                break
    print(f"\nTipos geometría: {dict(geom_types)}")
    if regions:
        print(f"Top regiones: {regions.most_common(10)}")

    # Guardar GeoJSON (visor)
    fc = {
        "type": "FeatureCollection",
        "_source": "Libro Geositios de Chile - SERNAGEOMIN 2024 (KMZ + XLSX oficial)",
        "_count": len(features),
        "features": features,
    }
    OUT_GEOJSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_GEOJSON.write_text(json.dumps(fc, ensure_ascii=False), encoding="utf-8")
    print(f"\n✓ GeoJSON: {OUT_GEOJSON} ({OUT_GEOJSON.stat().st_size/1024/1024:.2f} MB)")

    # Guardar JSON tabular para BD
    db = {
        "_source": "SERNAGEOMIN 2024 Libro Geositios de Chile",
        "_count": len(features),
        "geositios": [
            {
                "nombre": f["properties"]["nombre"],
                "descripcion": f["properties"].get("descripcion", "")[:500],
                "tipo_geometria": f["geometry"]["type"],
                "coordenadas": (f["geometry"]["coordinates"] if f["geometry"]["type"] == "Point"
                                else f["geometry"]["coordinates"][0][0] if f["geometry"]["coordinates"] else None),
                **{k: v for k, v in f["properties"].items() if k not in ("nombre", "descripcion")},
            }
            for f in features
        ],
    }
    OUT_DB.parent.mkdir(parents=True, exist_ok=True)
    OUT_DB.write_text(json.dumps(db, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"✓ DB JSON: {OUT_DB} ({OUT_DB.stat().st_size/1024:.1f} KB)")


if __name__ == "__main__":
    main()
